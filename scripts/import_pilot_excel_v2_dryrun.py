"""Dry-run v2: import MetaBreath Pilot Test 14-16 July.xlsx (Days 1-3).

Differences from v1 (scripts/import_pilot_excel_dryrun.py):
  - New Excel path (contains Days 1, 2, 3).
  - Skips separator rows ("Day 1", "Day 2", "Day 3", blank rows).
  - Parses text-format timestamps from Day 2/3 ("15/7/2026, 7:58:22").
  - Day 1 rows are idempotent via ON CONFLICT (time, device_id) DO NOTHING.
  - Force-overrides `man` profile to weight=53, height=172 (Day 1 was wrong).

Output goes to stdout; pipe to a .sql file for review before running.
"""
import openpyxl
from datetime import datetime, timedelta
from uuid import uuid4
import json

EXCEL_PATH = "/Users/ciy_th/Downloads/MetaBreath Pilot Test 14-16 July.xlsx"

USER_MAP = {
    "test1":     ("test1",        "1d8b7ffd-1ae8-4ecb-9982-c4b364268e6f"),
    "เบญจมาศ":  ("benjamars",    "1d605079-f882-4408-93b2-ff63608fca76"),
    "benjamars": ("benjamars",    "1d605079-f882-4408-93b2-ff63608fca76"),
    "wan":       ("Wan",          "7aaa453c-a8d2-4cf2-b61b-c8cdb3894d81"),
    "toom":      ("toom",         "bdaeb6d8-b877-49ba-9f3a-556bbfa09800"),
    "man":       ("man",          "ed06c0bb-598c-43a9-8439-04b5b9da5796"),
    "aor":       ("aor",          "5f774478-267b-4f16-9c77-efe45aee5b29"),
    "pui":       ("pui",          "ee1eee64-a65a-4a5f-893b-4b49e8ba335d"),
    "araya":     ("araya",        "5cebe88b-3f66-4987-98d0-4b328a8ec026"),
    "mark":      ("mark",         "6dee7753-311a-469c-9086-f8d1265b2c5f"),
    "orasinee":  ("orasinee",     "fa35f4bd-e360-4aa9-8d25-7c734253c64f"),
}

# Users whose profile weight/height should be force-updated (overrides even
# when the current value is non-NULL). Day 1 self-report was wrong.
FORCE_PROFILE_OVERRIDE = {
    "man": {"weight_kg": 53, "height_cm": 172},
}

SHARED_DEVICE_ID = "a1fe4813-ff62-4efb-8f29-af5c7d4f824e"

URINE_LEVEL_MAP = {
    "ระดับ 1 สีนู้ด":      ("trace",     0.5, 5.0),
    "ระดับ 2 สีส้มอ่อน":   ("small",     1.5, 15.0),
    "ระดับ 3":            ("moderate",  4.0, 40.0),
    "ระดับ 4":            ("large",     8.0, 80.0),
}

PPM_TO_MV = 10.0
SEPARATOR_TOKENS = {"day 1", "day 2", "day 3"}


def esc(s):
    if s is None:
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def sql_datetime(dt):
    if dt is None:
        return "NULL"
    return "'" + dt.strftime("%Y-%m-%d %H:%M:%S.%f") + "'"


def parse_ts(v):
    """Accept datetime or 'D/M/YYYY, H:M:S' text (Day 2/3 rows)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y, %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def process_row(row_dict, force_overrides_emitted):
    excel_username = (row_dict["username"] or "").strip()
    key = excel_username.lower()
    match = USER_MAP.get(key) or USER_MAP.get(excel_username)
    if match is None:
        yield f"-- SKIP: no user match for '{excel_username}'"
        return

    db_username, user_id = match
    ts = row_dict["timestamp"]
    if ts is None:
        yield f"-- SKIP {excel_username}: unparseable timestamp"
        return
    session_id = f"{db_username.lower()}-pilot-{ts.strftime('%Y%m%d')}"

    yield f"\n-- ═══════════════════════════════════════════════════════════"
    yield f"-- {ts.strftime('%Y-%m-%d %H:%M')} | {excel_username} → {db_username}"
    yield f"-- ═══════════════════════════════════════════════════════════"

    # ── 1) Profile weight/height ──
    if db_username.lower() in FORCE_PROFILE_OVERRIDE and db_username.lower() not in force_overrides_emitted:
        ov = FORCE_PROFILE_OVERRIDE[db_username.lower()]
        set_pairs = [f"{k} = {v}" for k, v in ov.items()]
        yield (
            f"-- FORCE override profile (Day 1 self-report was wrong)\n"
            f"UPDATE profiles SET {', '.join(set_pairs)}\n"
            f"  WHERE user_id = '{user_id}';"
        )
        force_overrides_emitted.add(db_username.lower())
    else:
        weight = row_dict.get("weight_kg")
        height = row_dict.get("height_cm")
        if (weight is not None or height is not None) and db_username.lower() not in FORCE_PROFILE_OVERRIDE:
            set_pairs = []
            null_conds = []
            if weight is not None:
                set_pairs.append(f"weight_kg = {weight}")
                null_conds.append("weight_kg IS NULL")
            if height is not None:
                set_pairs.append(f"height_cm = {height}")
                null_conds.append("height_cm IS NULL")
            yield (
                f"UPDATE profiles SET {', '.join(set_pairs)}\n"
                f"  WHERE user_id = '{user_id}'\n"
                f"    AND ({' OR '.join(null_conds)});"
            )

    # ── 2) Insert 3 breath readings as one pilot session ──
    ppms = [row_dict.get("ppm1"), row_dict.get("ppm2"), row_dict.get("ppm3")]
    ppms = [p for p in ppms if p is not None]
    if ppms:
        q = {
            "source":              f"excel_pilot_import_{ts.strftime('%Y_%m_%d')}",
            "age_range":           row_dict.get("age_range"),
            "last_meal_time":      row_dict.get("last_meal_time"),
            "last_meal_type":      row_dict.get("last_meal_type"),
            "alcohol_24h":         row_dict.get("alcohol"),
            "sleep_hours":         row_dict.get("sleep"),
            "exercise_intensity":  row_dict.get("exercise"),
            "stress_today":        row_dict.get("stress"),
            "medication":          row_dict.get("medication"),
            "smoke_2h":            row_dict.get("smoke"),
            "brush_30m":           row_dict.get("brush"),
            "urine_level_note":    row_dict.get("urine_note"),
        }
        q_json = json.dumps({k: v for k, v in q.items() if v is not None}, ensure_ascii=False)

        for i, ppm in enumerate(ppms):
            reading_time = ts + timedelta(seconds=60 * i)
            mv = round(ppm * PPM_TO_MV, 4)
            raw_json = q_json if i == 0 else "{}"
            yield (
                f"INSERT INTO sensor_readings\n"
                f"  (time, device_id, user_id, session_id,\n"
                f"   ambient_voc, breath_voc, acetone_delta,\n"
                f"   quality_score, reliability_score, environment_penalty,\n"
                f"   confidence_score, label, raw)\n"
                f"VALUES\n"
                f"  ({sql_datetime(reading_time)}, '{SHARED_DEVICE_ID}', '{user_id}',\n"
                f"   {esc(session_id)},\n"
                f"   0.7, {round(0.7 + mv/1000.0, 6)}, {mv},\n"
                f"   100, 95, 0.05,\n"
                f"   0.95, 'low', {esc(raw_json)}::jsonb)\n"
                f"ON CONFLICT (time, device_id) DO NOTHING;"
            )

    # ── 3) Urine ketone log ──
    urine_note = row_dict.get("urine_note")
    if urine_note:
        cat_key = None
        for k in URINE_LEVEL_MAP:
            if urine_note.strip().startswith(k[:8]):
                cat_key = k
                break
        if cat_key:
            category, mmol, mgdl = URINE_LEVEL_MAP[cat_key]
            log_id = str(uuid4())
            # No unique index on (user_id, ts, source) — guard with NOT EXISTS
            # so re-running the script doesn't duplicate Day 1 rows.
            yield (
                f"INSERT INTO ketone_logs\n"
                f"  (id, user_id, ts, value_mmol, source, note,\n"
                f"   ketone_type, urine_category, urine_mg_dl)\n"
                f"SELECT '{log_id}', '{user_id}', {sql_datetime(ts)}, {mmol},\n"
                f"       'excel_pilot_import', {esc(f'imported: {urine_note}')},\n"
                f"       'urine', {esc(category)}, {mgdl}\n"
                f"WHERE NOT EXISTS (\n"
                f"  SELECT 1 FROM ketone_logs\n"
                f"   WHERE user_id = '{user_id}'\n"
                f"     AND ts = {sql_datetime(ts)}\n"
                f"     AND source = 'excel_pilot_import'\n"
                f");"
            )
        else:
            yield f"-- SKIP urine: unrecognised note '{urine_note}'"


def is_separator(row):
    """Row like ('Day 1 ', None, None, ...) or all-blank/whitespace."""
    non_null = [c for c in row if c is not None]
    if not non_null:
        return True
    if len(non_null) == 1:
        s = str(non_null[0]).strip().lower()
        if not s or s in SEPARATOR_TOKENS:
            return True
    return False


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    def get(row, key):
        return row[idx[key]] if key in idx else None

    print("-- ═══════════════════════════════════════════════════════════")
    print("-- MetaBreath Pilot Test 14-16 July — DRY RUN v2")
    print("-- Generated by scripts/import_pilot_excel_v2_dryrun.py")
    print(f"-- Source: {EXCEL_PATH}")
    print("-- Day 1 inserts are idempotent (ON CONFLICT (time,device_id) DO NOTHING)")
    print("-- ═══════════════════════════════════════════════════════════")
    print("BEGIN;")

    force_overrides_emitted = set()
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if is_separator(row):
            continue
        username_cell = get(row, "1. ชื่อ (Username ที่ใช้ในแอพลิเคชั่น MetaBreath)")
        if not username_cell:
            continue

        row_dict = {
            "timestamp":      parse_ts(get(row, "ประทับเวลา")),
            "username":       username_cell,
            "ppm1":           get(row, "ppm 1"),
            "ppm2":           get(row, "ppm 2"),
            "ppm3":           get(row, "ppm 3"),
            "age_range":      get(row, "2. ช่วงอายุ"),
            "weight_kg":      get(row, "3. น้ำหนัก (กิโลกรัม)"),
            "height_cm":      get(row, "4. ส่วนสูง (เซ็นติเมตร)"),
            "last_meal_time": get(row, "5. ช่วงเวลาที่กินอาหารมื้อล่าสุด"),
            "last_meal_type": get(row, "6. ลักษณะอาหารมื้อล่าสุด (เลือกได้หลายข้อ)"),
            "alcohol":        get(row, "7.  ดื่มแอลกอฮอล์ใน 24 ชั่วโมงที่ผ่านมาหรือไม่ "),
            "sleep":          get(row, "8. ชั่วโมงการนอนคืนที่ผ่านมา "),
            "exercise":       get(row, "9. ระดับความหนักของการออกกำลังกายใน 24 ชั่วโมงที่ผ่านมา"),
            "stress":         get(row, "10. ระดับความเครียดวันนี้ "),
            "medication":     get(row, "11.  กำลังใช้ยาที่อาจมีผลต่อเมตาบอลิซึมหรือไม่ (เช่น ยาเบาหวาน, ยาลดไขมัน, อินซูลิน, สเตียรอยด์, ยาหรือฮอร์โมนไทรอยด์) "),
            "smoke":          get(row, "12.  สูบบุหรี่/บุหรี่ไฟฟ้าใน 2 ชั่วโมงก่อนวัดหรือไม่\n"),
            "brush":          get(row, "13.  แปรงฟัน/ใช้น้ำยาบ้วนปากใน 30 นาทีก่อนวัดหรือไม่ "),
            "urine_note":     get(row, "14. ระบุแถบสีจากการตรวจคีโตนในปัสสาวะเช้านี้"),
        }
        for line in process_row(row_dict, force_overrides_emitted):
            print(line)
        count += 1

    print("\n-- ═══════════════════════════════════════════════════════════")
    print(f"-- Processed {count} data rows")
    print("-- ═══════════════════════════════════════════════════════════")
    print("-- Review, then change ROLLBACK to COMMIT to apply.")
    print("ROLLBACK;")


if __name__ == "__main__":
    main()
